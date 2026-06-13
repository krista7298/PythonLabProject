from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference

def convertData(tempF):
    return (tempF - 32) * 5 / 9


def insertData(path, data):

    try:
        with open(path, "a") as file:
            file.write(data + "\n")

    except Exception as e:
        print("Error writing to file:", e)


def viewData(path):

    try:
        print("The file", path)

        with open(path, "r") as file:
            contents = file.read()
            print(contents)

    except Exception as e:
        print("Error reading file:", e)


def getInput():

    entryCount = int(input("How many entries are you inputting?\n"))

    for counter in range(entryCount):

        dateEntered = input("\nEnter a date:\n")

        tempF = float(input("Enter the highest temp for the inputted date:\n"))

        convertedTemp = convertData(tempF)

        data = f"{dateEntered},{tempF},{convertedTemp}"

        try:

            insertData("ZooData.csv", data)

            print(f"The following data was saved at {datetime.now()} :")
            print(data)

        except Exception as e:
            print("Error saving data:", e)


# Function: createChart
# Argument 1: path to CSV file (string)
# Argument 2: chart type (string)
# Return Value: None
def createChart(path, chartType):

    print("\nChoose Data Source")
    print("1 Fahrenheit")
    print("2 Celsius")

    sourceChoice = input("Enter selection: ")

    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Date"

    if sourceChoice == "1":
        ws["B1"] = "Fahrenheit"
    else:
        ws["B1"] = "Celsius"

    rowNumber = 2

    with open(path, "r") as file:

        for line in file:

            data = line.strip().split(",")

            if len(data) >= 3:

                ws.cell(row=rowNumber, column=1).value = data[0]

                if sourceChoice == "1":
                    ws.cell(row=rowNumber, column=2).value = float(data[1])
                else:
                    ws.cell(row=rowNumber, column=2).value = float(data[2])

                rowNumber += 1

    labels = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=rowNumber - 1
    )

    chartData = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=rowNumber - 1
    )

    if chartType == "bar":
        chart = BarChart()
    else:
        chart = LineChart()

    chart.title = "kriagu7298 06/12/2026"

    chart.x_axis.title = "Date"

    if sourceChoice == "1":
        chart.y_axis.title = "Fahrenheit"
    else:
        chart.y_axis.title = "Celsius"

    chart.add_data(chartData, titles_from_data=True)

    chart.set_categories(labels)

    ws.add_chart(chart, "D2")

    wb.save("final.xlsx")

    print("Chart saved as final.xlsx")


# Function: generateReport
# Argument: path to CSV file (string)
# Return Value: None
def generateReport(path):

    print("\nChoose Graph Type")
    print("1 Bar Chart")
    print("2 Line Chart")

    graphChoice = input("Enter selection: ")

    if graphChoice == "1":
        createChart(path, "bar")
    else:
        createChart(path, "line")


print("kriagu7298 Spreadsheet Automation Menu")

menuOptions = [
    "1 Input Data",
    "2 View Current Data",
    "3 Generate Report"
]

print("Choose a number from the following options")

for option in menuOptions:
    print(option)

user_choice = input()

if user_choice == "1":

    print("You selected", user_choice, "at", str(datetime.now()))
    getInput()

elif user_choice == "2":

    print("You selected", user_choice, "at", str(datetime.now()))
    viewData("ZooData.csv")

elif user_choice == "3":

    print("You selected", user_choice, "at", str(datetime.now()))
    generateReport("ZooData.csv")

else:

    print("Invalid Selection")
